import re
from openpyxl import load_workbook
from utils.helpers import (
    get_header_row, unwrap_merged_headers, copy_table, copy_metadata
)

SOURCE_FILE = "inputs/samples/Mapping_1-7.xlsx"
TARGET_FILE = "templates/mapping_1-11.xlsx"
OUTPUT_FILE = "outputs/mapping_1-11_UPDATED.xlsx"
TEMPLATE_SHEET_NAME = "Transformation - Sourcing (1)"


def run():
    wb_src = load_workbook(SOURCE_FILE, data_only=False)
    wb_tgt = load_workbook(TARGET_FILE)

    pattern = re.compile(r"^Transformation - Sourcing", re.IGNORECASE)
    src_sheet_names = [name for name in wb_src.sheetnames if pattern.match(name)]

    for sheet_name in src_sheet_names:
        print(f"\n🔁 Processing: {sheet_name}")
        ws_src = wb_src[sheet_name]

        # Create target sheet if needed
        if sheet_name not in wb_tgt.sheetnames:
            print(f"📄 Creating target sheet: '{sheet_name}' from template...")
            if TEMPLATE_SHEET_NAME not in wb_tgt.sheetnames:
                raise Exception(f"❌ Template sheet '{TEMPLATE_SHEET_NAME}' not found in target.")
            template_ws = wb_tgt[TEMPLATE_SHEET_NAME]
            copied_ws = wb_tgt.copy_worksheet(template_ws)
            copied_ws.title = sheet_name

        ws_tgt = wb_tgt[sheet_name]

        # ✅ Copy metadata top rows like Tablename, Description, etc.
        copy_metadata(ws_src, ws_tgt)

        # ✅ Step 1: unwrap all potential header rows BEFORE header detection
        for r in range(1, 30):  # Unwrap possible header rows
            unwrap_merged_headers(ws_src, r)
            unwrap_merged_headers(ws_tgt, r)

        # ✅ Step 2: Find header rows after unmerging
        src_row_sourcing = get_header_row(ws_src, "Dependency")
        tgt_row_sourcing = get_header_row(ws_tgt, "Dependency")
        src_row_transform = get_header_row(ws_src, "#")
        tgt_row_transform = get_header_row(ws_tgt, "#")

        print(f"  📌 Header rows - Sourcing: src={src_row_sourcing}, tgt={tgt_row_sourcing}")
        print(f"  📌 Header rows - Transformation: src={src_row_transform}, tgt={tgt_row_transform}")

        # Step 3: Temporarily unmerge target to prevent overlap issues
        merges = list(ws_tgt.merged_cells.ranges)
        for rng in merges:
            ws_tgt.unmerge_cells(str(rng))

        # Step 4: Copy both tables
        copy_table(ws_src, ws_tgt, src_row_sourcing, tgt_row_sourcing, label="Sourcing")
        copy_table(ws_src, ws_tgt, src_row_transform, tgt_row_transform, label="Transformation")

        # Step 5: Re-merge
        for rng in merges:
            ws_tgt.merge_cells(str(rng))

    # Final save
    wb_tgt.save(OUTPUT_FILE)
    print(f"\n✅ All transformation sheets copied and saved to {OUTPUT_FILE}")



if __name__ == "__main__":
    run()
