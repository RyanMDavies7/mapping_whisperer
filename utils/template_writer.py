# utils/template_writer.py

"""
Writes one entity’s data into a fresh copy of the
Mapping Specification Excel template—preserving all
formatting, merges, and conditional format rules.

Outputs under 'outputs/' with the file name <EntityName>.xlsx.
"""

from pathlib import Path
from openpyxl import load_workbook
from utils.helpers import get_header_row, unwrap_merged_headers, normalize
import copy

# Map of our internal keys → exact Excel header captions
HDRS = {
  "name":"Attribute Name","desc":"Attribute Description","type":"Datatype",
  "sd":"Sourced/Derived","src_t":"Source Table","src_a":"Source Attribute",
  "ref_dim":"Referenced Dimension","nn":"Not Null",
  "def_v":"Default Values","def_m1":"Default Records",
  "def_m2":"Default Records (2)","clust":"Clustering","part":"Partitioning",
  "key":"Keys"
}

def _build_map(ws, hdr_row: int) -> dict[str,int]:
    # Unmerge and unwrap any merged headers first
    unwrap_merged_headers(ws, hdr_row)
    return { normalize(c.value): c.column for c in ws[hdr_row] if c.value }

def _copy_style_only(ws, src_row: int, dst_row: int) -> None:
    # Clone everything BUT the .value
    for col in range(1, ws.max_column+1):
        s, d = ws.cell(src_row,col), ws.cell(dst_row,col)
        if s.has_style:
            d._style = copy.copy(s._style)
        d.number_format = s.number_format

def _write_sourcing(ws, entity: dict) -> None:
    """Populate the Sourcing table from entity['sources']"""
    hdr = get_header_row(ws, "Dependency")
    colmap = { normalize(c.value): c.column for c in ws[hdr] if c.value }
    for i, src in enumerate(entity.get("sources", [])):
        r = hdr + 1 + i
        if "source database" in colmap:
            ws.cell(r, colmap["source database"]).value = src["database"]
        if "table name/dataset name" in colmap:
            ws.cell(r, colmap["table name/dataset name"]).value = src["table"]
        if "source column" in colmap:
            ws.cell(r, colmap["source column"]).value = src["column"]

def write_entity(entity: dict, template_path: str|Path, out_dir: str|Path) -> Path:
    """
    Write the entity to a new workbook:
      1) Fill metadata (B3/B4)
      2) Table options (above # header)
      3) Sourcing table
      4) Transformation table
      5) Preserve all styles + merges
    """
    wb = load_workbook(template_path)
    ws = wb["Transformation - Sourcing (1)"]

    # --- 1) Top metadata ---
    ws["B3"] = entity["name"]
    ws["B4"] = entity["description"]

    # --- 2) Table options row ---
    hdr = get_header_row(ws, "#")
    opts = entity.get("table_options","")
    if opts:
        ws.cell(hdr-1, 1).value = f"Table options: {opts}"

    # --- 3) Sourcing ---
    _write_sourcing(ws, entity)

    # --- 4) Transformation headers & mapping ---
    trans_hdr = hdr
    cmap      = _build_map(ws, trans_hdr)

    # Validate headers exist
    for key, caption in HDRS.items():
        if normalize(caption) not in cmap:
            raise ValueError(f"Missing header: {caption}")

    start = trans_hdr + 1
    needed = len(entity["fields"])
    CAP = 500

    # Expand if >500 pre-seeded rows
    if needed > CAP:
        extra = needed - CAP
        ins   = start + CAP
        ws.insert_rows(ins, extra)
        for off in range(extra):
            _copy_style_only(ws, start+CAP-1, ins+off)

    # Ensure style present on every row
    for i in range(needed):
        _copy_style_only(ws, start, start+i)

    # Helper to lookup column by key
    def C(k): return cmap[normalize(HDRS[k])]

    # --- 5) Write each field row (skip '#' col) ---
    for i, f in enumerate(entity["fields"]):
        r = start + i
        ws.cell(r,C("name")).value    = f["name"]
        ws.cell(r,C("desc")).value    = f["description"]
        ws.cell(r,C("type")).value    = f["datatype"]
        ws.cell(r,C("sd")).value      = "Sourced" if f["sourced"] else "Derived"
        ws.cell(r,C("src_t")).value   = f["src_table"]
        ws.cell(r,C("src_a")).value   = f["src_attr"]
        ws.cell(r,C("ref_dim")).value = f["referenced_dimension"]
        ws.cell(r,C("nn")).value      = "Y" if f["not_null"] else "N"
        ws.cell(r,C("def_v")).value   = f["def_val"]
        ws.cell(r,C("def_m1")).value  = f["def_m1"]
        ws.cell(r,C("def_m2")).value  = f["def_m2"]
        ws.cell(r,C("clust")).value   = f["clustering"]
        ws.cell(r,C("part")).value    = f["partitioning"]
        ws.cell(r,C("key")).value     = f["key_type"]

    out = Path(out_dir)/f"{entity['name']}.xlsx"
    wb.save(out)
    return out
